import os
import json
import datetime
import uuid
from collections import defaultdict
import requests
from slugify import slugify

from flask import Flask, render_template, request, jsonify, redirect, url_for
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import tempfile

app = Flask(__name__)

# Load env vars
FB_ACCESS_TOKEN = os.getenv('FBACCSESSTOKEN', 'your_access_token_here')
ADMIN_TOKEN = os.getenv('ADMIN_TOKEN', 'admin123')

BUSINESS_CATALOGS = {
    'CASTELPHARMA': '1341574140528980',  # castel pharma
    'FOFO': '1816910538949001',  # fofo
    'SUDIID': '1767558300630702',  # Azucar sudi
    'UNILEVERID': '1605226514252142',  # argento UNILEVER
}

# Global city_area dict
CITY_AREA_DICT = {}

# Shipping prices by city
SHIPPING_PRICES = {
    'AinShams': 65,
    'Al-agamy': 75,
    'Alex': 75,
    'Aswan': 130,
    'Asyut': 95,
    'Badrashin': 75,
    'Abu Tesht': 75,
    'Farshut': 75,
    'Dar El-Salam': 65,
    'Al Balena': 95,
    'Gerga': 95,
    'Banha': 75,
    'Behira': 75,
    'BeniSuef': 95,
    'Helwan': 65,
    'Damietta': 75,
    'Dekernes': 75,
    'Desouk': 75,
    'Dokki': 65,
    'Downtown': 65,
    'Abu Sinbil': 130,
    'Administrative Capital': 65,
    'New Heliopolis City': 65,
    'Cairo Governorate Desert': 65,
    'Zaafarana': 130,
    'Shalateen': 130,
    'Marsa Alam': 130,
    'Halaib': 130,
    'Sallum': 130,
    'Siwa Oasis': 130,
    'Sidi Barrani': 130,
    'Farafra': 130,
    'Dakhla': 75,
    'Kharga': 130,
    'Sharm El-Sheikh': 130,
    'Abu Radis': 130,
    'Ain Sokhna': 85,
    'Faiyum': 95,
    'Faqus': 75,
    'Fardos': 65,
    'Faisal': 65,
    'Haram': 65,
    'Hurghada': 130,
    'Ismailia': 85,
    'Kafr El-Sheikh': 75,
    'Khanka': 65,
    'Luxor': 130,
    'Maadi': 65,
    'Mahala': 75,
    'Mansoura': 75,
    'Matamir': 65,
    'Dayrout': 95,
    'Al Qusiyyah': 130,
    'Malawi': 130,
    'Dayr Mawas': 130,
    'Menya': 95,
    'Moharram Bek': 75,
    'Monufia': 75,
    'Nasr City': 65,
    'New Cairo': 65,
    'Matrouh': 130,
    'North Coast': 130,
    'Oct6th': 65,
    'Port said': 85,
    'Qena': 130,
    'Sharqia': 75,
    'Shorouk': 65,
    'Shoubara El Khima': 65,
    'Sohag': 95,
    'Suez': 85,
    'Tanash': 65,
    'Tanta': 75,
    'Zayton': 65,
    '10th of Ramadan City': 75
}

# Fixed sender data for each catalog
CATALOG_SENDERS = {
    'SUDIID': {
        'name': 'Azucar Sharqia',
        'city': 'Zagazig',
        'area': 'حي الزهور',
        'address': 'حي الزهور',
        'phone': '+20 10 17549330',
        'email': ''
    },
    'UNILEVERID': {
        'name': 'argento Sharqia',
        'city': 'Zagazig',
        'area': 'حي الزهور',
        'address': 'حي الزهور',
        'phone': '01055688136',
        'email': ''
    },
    'FOFO': {
        'name': 'fofo Sharqia',
        'city': 'Zagazig',
        'area': 'شارع العصلوجي',
        'address': 'شارع العصلوجي',
        'phone': '+20 12 12137256',
        'email': ''
    },
    'CASTELPHARMA': {
        'name': 'castel pharma Sharqia',
        'city': 'Zagazig',
        'area': 'شارع فلل الجامعه',
        'address': 'شارع فلل الجامعه',
        'phone': '',
        'email': ''
    }
}

def load_city_area_dict():
    """Load city to areas mapping from addresses.xlsx"""
    global CITY_AREA_DICT
    CITY_AREA_DICT = defaultdict(list)
    if os.path.exists('addresses.xlsx'):
        try:
            wb = load_workbook('addresses.xlsx')
            if 'Speedaf standard address data' in wb.sheetnames:
                sheet = wb['Speedaf standard address data']
                header = [cell.value for cell in sheet[1] if cell.value]
                print(f"Header found: {header}")
                if 'City' in header and 'Area' in header:
                    city_idx = header.index('City')
                    area_idx = header.index('Area')
                    for row in sheet.iter_rows(min_row=2):
                        city = row[city_idx].value
                        area = row[area_idx].value
                        if city and area:
                            city_key = str(city).strip()
                            area_value = str(area).strip()
                            if area_value not in CITY_AREA_DICT[city_key]:
                                CITY_AREA_DICT[city_key].append(area_value)
                else:
                    print("City or Area column not found in header")
            else:
                print("Sheet 'Speedaf standard address data' not found")
        except Exception as e:
            print(f"Error loading addresses.xlsx: {e}")
    else:
        print("addresses.xlsx file not found")
    
    print(f"Loaded {len(CITY_AREA_DICT)} cities with areas")
    # Debug: print first few cities
    if CITY_AREA_DICT:
        print("Sample cities:", list(CITY_AREA_DICT.keys())[:5])

def load_catalog():
    """Load products from first valid catalog file"""
    paths = ['data/catalog_cache.json', 'catalog.json', 'data/catalogs/latest.json']
    for path in paths:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data, path
                    elif isinstance(data, dict) and 'products' in data:
                        return data['products'], path
            except Exception as e:
                print(f"Error loading {path}: {e}")
    return [], None

def save_catalog_products(products, out_path=None):
    """Save products to JSON safely"""
    if out_path is None:
        out_path = 'data/catalog_cache.json'
    dir_path = os.path.dirname(out_path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)
    temp_path = out_path + '.tmp'
    try:
        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(products, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, out_path)
        print(f"Saved {len(products)} products to {out_path}")
    except Exception as e:
        print(f"Error saving {out_path}: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)

def fetch_facebook_products(catalog_id, access_token):
    """Fetch products from Facebook catalog"""
    url = f"https://graph.facebook.com/v18.0/{catalog_id}/products"
    params = {
        'fields': 'id,name,description,price,currency,availability,brand,image_url',
        'access_token': access_token
    }
    response = requests.get(url, params=params)
    print(f"Response for {catalog_id}: {response.status_code}")
    if response.status_code == 200:
        data = response.json()
        print(f"Raw data from {catalog_id}: {data}")
        all_products = data.get('data', [])
        available_products = [p for p in all_products if p.get('availability') in ['available', 'in stock', None]]
        print(f"Total products: {len(all_products)}, Available: {len(available_products)}")
        return available_products
    else:
        print(f"Failed to fetch {catalog_id}: {response.status_code} - {response.text}")
        return []

def update_catalogs():
    """Fetch and merge products without duplicates"""
    all_products = []
    seen_ids = set()
    for name, cid in BUSINESS_CATALOGS.items():
        prods = fetch_facebook_products(cid, FB_ACCESS_TOKEN)
        print(f"Fetched {len(prods)} from {name}")
        for p in prods:
            pid = p.get('id')
            if pid not in seen_ids:
                seen_ids.add(pid)
                # Handle price parsing - Facebook returns price as string like "EGP550.00"
                price_str = p.get('price', '0')
                if isinstance(price_str, str):
                    # Extract numeric value from string like "EGP550.00" or "EGP1,970.00"
                    import re
                    price_match = re.search(r'[\d,]+\.?\d*', price_str.replace(',', ''))
                    price_value = float(price_match.group()) if price_match else 0.0
                    # Extract currency from string
                    currency_match = re.search(r'[A-Z]{3}', price_str)
                    currency = currency_match.group() if currency_match else 'EGP'
                else:
                    price_value = 0.0
                    currency = 'EGP'
                
                all_products.append({
                    'id': pid,
                    'sku': f"{name}-{pid}",
                    'title': p.get('name', ''),
                    'description': p.get('description', ''),
                    'price': price_value,
                    'currency': currency,
                    'brand': p.get('brand', ''),
                    'image_url': p.get('image_url', ''),
                    'shipping_price': 50.0,  # default
                    'free_shipping': False,
                    'offers': []
                })
    return all_products

def generate_landing_links(products, write_back=True):
    """Add website slug to products if missing"""
    changed = 0
    for p in products:
        if 'website' not in p or not p['website']:
            slug = p.get('id') or p.get('sku') or slugify(p.get('title', ''))
            p['website'] = f"/landing/{slug}"
            changed += 1
    if write_back and changed > 0:
        save_catalog_products(products)
    return changed, 'data/catalog_cache.json'

def find_product_by_slug(slug):
    """Find product by slug in website, id, sku, or title"""
    products, _ = load_catalog()
    for p in products:
        if p.get('website', '').endswith(f"/{slug}"):
            return p
        if str(p.get('id', '')) == slug or str(p.get('sku', '')) == slug:
            return p
        if slugify(p.get('title', '')) == slug:
            return p
    return None

def calculate_order(product, quantity, customer_city=None, offer=None):
    """Calculate prices with offers and shipping"""
    price_per = product['price']
    subtotal = price_per * quantity
    
    # Calculate shipping based on customer city
    if product.get('free_shipping'):
        shipping_applied = 0
    elif customer_city and customer_city in SHIPPING_PRICES:
        shipping_applied = SHIPPING_PRICES[customer_city]
    else:
        # Fallback to product's default shipping price
        shipping_applied = product.get('shipping_price', 50.0)
    
    discount = 0
    offer = product.get('offers', [])[-1] if not offer else offer  # latest if none specified
    if offer:
        if offer['type'] == 'percentage':
            discount = subtotal * (offer['value'] / 100)
        elif offer['type'] == 'fixed':
            discount = min(offer['value'], subtotal)
    total = subtotal - discount + shipping_applied
    return subtotal, discount, shipping_applied, total

def add_order_to_excel(order):
    """Append order to addresses.xlsx sheet '0'"""
    if not os.path.exists('addresses.xlsx'):
        wb = Workbook()
        wb['Sheet'].title = '0'
        wb.save('addresses.xlsx')

    def atomic_write(data):
        with tempfile.NamedTemporaryFile(mode='w+', suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        try:
            orig_wb = load_workbook('addresses.xlsx')
            if '0' not in orig_wb.sheetnames:
                orig_wb.create_sheet('0')
            sheet = orig_wb['0']
            row = len(sheet['A']) + 1
            for i, val in enumerate(data, 1):
                sheet.cell(row=row, column=i, value=val)
            orig_wb.save(temp_path)
            os.replace(temp_path, 'addresses.xlsx')
        except Exception as e:
            print(f"Error adding to Excel: {e}")
            if os.path.exists(temp_path):
                os.remove(temp_path)

    # Map order to Excel columns
    # Assuming columns as in read_file: S.O., Goods type, ..., Receiver Email, Delivery Type
    goods = {'name': order['product']['title'], 'quantity': order['quantity'], 'price': order['price']}
    sender = order.get('customer', {}).get('sender', {})
    receiver = order['customer']
    data = [
        '',  # S.O.
        'product',  # Goods type
        order['product']['title'],  # Goods name
        order['quantity'],  # Quantity
        0.5,  # Weight
        order['total'],  # COD
        0,  # Insure price
        'No',  # Whether to allow the package to be opened
        order.get('notes', ''),  # Remark
        sender.get('name', ''), sender.get('phone', ''), sender.get('city', ''), sender.get('area', ''), sender.get('address', ''), sender.get('email', ''),
        receiver.get('name', ''), receiver.get('phone', ''), receiver.get('city', ''), receiver.get('area', ''), receiver.get('address', ''), receiver.get('email', ''),
        'standard'  # Delivery Type
    ]
    atomic_write(data)

def archive_and_reset_orders(archive_dir='data/archives'):
    """Archive current addresses.xlsx to data/archives with timestamp and reset sheet '0'"""
    if not os.path.exists('addresses.xlsx'):
        return False
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(archive_dir, exist_ok=True)
    archive_path = os.path.join(archive_dir, f'orders_{timestamp}.xlsx')
    try:
        import shutil
        shutil.copy('addresses.xlsx', archive_path)
        # Reset sheet '0'
        wb = load_workbook('addresses.xlsx')
        if '0' in wb.sheetnames:
            wb['0'].delete_rows(2, wb['0'].max_row)  # keeps header
            wb.save('addresses.xlsx')
        return True, archive_path
    except Exception as e:
        print(f"Archive error: {e}")
        return False, None

# Routes
@app.route('/')
def index():
    products, _ = load_catalog()
    if not products:
        # Fallback to first product if list
        return render_template('index.html', products=[])
    return render_template('index.html', products=products)

@app.route('/api/products')
def api_products():
    products, _ = load_catalog()
    return jsonify({'ok': True, 'count': len(products), 'products': products})

@app.route('/landing/<slug>')
def product_landing(slug):
    product = find_product_by_slug(slug)
    print(f"Product found: {product is not None}")
    print(f"City areas dict size: {len(CITY_AREA_DICT)}")
    print(f"Sample city areas: {dict(list(CITY_AREA_DICT.items())[:3])}")
    
    if product:
        return render_template('product_landing.html', product=product, city_areas=dict(CITY_AREA_DICT))
    # Fallback
    products, _ = load_catalog()
    if products:
        return render_template('product_landing.html', product=products[0], city_areas=dict(CITY_AREA_DICT))
    return render_template('index.html', products=[])

@app.route('/api/landing_order', methods=['POST'])
def landing_order():
    try:
        # Parse JSON data
        if request.is_json:
            data = request.get_json()
        else:
            # Try to parse form data as JSON
            data = request.form.to_dict()
            if 'customer' in data and isinstance(data['customer'], str):
                import ast
                try:
                    data['customer'] = ast.literal_eval(data['customer'])
                except:
                    pass

        product_id = data.get('product_id')
        quantity = int(data.get('quantity', 1))
        customer = data.get('customer', {})

        # Validate required fields
        if not product_id:
            return jsonify({'error': 'معرف المنتج مطلوب'}), 400
        
        if not customer.get('name'):
            return jsonify({'error': 'اسم المستلم مطلوب'}), 400
        
        if not customer.get('phone'):
            return jsonify({'error': 'هاتف المستلم مطلوب'}), 400

        products, _ = load_catalog()
        product = next((p for p in products if str(p.get('id', '')) == str(product_id)), None)
        if not product:
            return jsonify({'error': 'المنتج غير موجود'}), 400

        customer_city = customer.get('city')
        subtotal, discount, shipping_applied, total = calculate_order(product, quantity, customer_city)

        order = {
            'id': str(uuid.uuid4()),
            'created_at': datetime.datetime.utcnow().isoformat(),
            'product': {k: v for k, v in product.items() if k in ['id', 'title', 'price']},
            'quantity': quantity,
            'subtotal': subtotal,
            'discount': discount,
            'shipping': shipping_applied,
            'total': total,
            'status': 'new',
            'customer': customer,
            'raw': data
        }

        # Load existing orders
        if os.path.exists('data/orders.json'):
            try:
                with open('data/orders.json', 'r', encoding='utf-8') as f:
                    orders = json.load(f)
                if not isinstance(orders, list):
                    orders = []
            except Exception as e:
                print(f"Error loading orders: {e}")
                orders = []
        else:
            orders = []

        # Insert at top
        orders.insert(0, order)

        # Save orders
        try:
            with open('data/orders.json', 'w', encoding='utf-8') as f:
                json.dump(orders, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving orders: {e}")
            return jsonify({'error': 'فشل حفظ الطلب'}), 500

        # Add to Excel (non-blocking)
        try:
            add_order_to_excel(order)
        except Exception as e:
            print(f"Error adding to Excel: {e}")
            # Don't fail the request if Excel fails

        return jsonify({'ok': True, 'order_id': order['id']})

    except Exception as e:
        print(f"Error in landing_order: {e}")
        return jsonify({'error': f'خطأ في المعالجة: {str(e)}'}), 500

@app.route('/dashboard')
def dashboard():
    token = request.args.get('token') or request.cookies.get('admin_token')
    if token != ADMIN_TOKEN:
        return jsonify({'error': 'Admin access denied'}), 403

    products, _ = load_catalog()
    if os.path.exists('data/orders.json'):
        try:
            with open('data/orders.json', 'r', encoding='utf-8') as f:
                orders = json.load(f)
        except:
            orders = []
    else:
        orders = []

    today = datetime.datetime.utcnow().date().isoformat()
    today_orders = [o for o in orders if o.get('created_at', '').startswith(today)]
    today_sales = sum(float(o.get('total', 0)) for o in today_orders)

    stats = {
        'total_orders': len(orders),
        'today_orders': len(today_orders),
        'today_sales': today_sales,
        'total_products': len(products),
        'free_shipping_products': len([p for p in products if p.get('free_shipping')])
    }

    return render_template('dashboard.html', stats=stats, products=products)

@app.route('/admin')
def admin():
    token = request.args.get('token') or request.cookies.get('admin_token')
    if token != ADMIN_TOKEN:
        return jsonify({'error': 'Admin access denied'}), 403

    products, _ = load_catalog()
    if os.path.exists('data/orders.json'):
        try:
            with open('data/orders.json', 'r', encoding='utf-8') as f:
                orders = json.load(f)
        except:
            orders = []
    else:
        orders = []

    today = datetime.datetime.utcnow().date().isoformat()
    today_orders = [o for o in orders if o.get('created_at', '').startswith(today)]
    today_sales = sum(float(o.get('total', 0)) for o in today_orders)

    stats = {
        'total_orders': len(orders),
        'today_orders': len(today_orders),
        'today_sales': today_sales,
        'total_products': len(products),
        'free_shipping_products': len([p for p in products if p.get('free_shipping')])
    }

    return render_template('admin.html', stats=stats, products=products, orders=orders[:50])  # last 50

@app.route('/api/update_catalog', methods=['POST'])
def api_update_catalog():
    if request.headers.get('Authorization') != f"Bearer {ADMIN_TOKEN}":
        return jsonify({'error': 'Unauthorized'}), 401

    new_products = update_catalogs()
    if new_products:
        # Merge with existing without duplicates based on id
        old_products, _ = load_catalog()
        seen_ids = {p.get('id'): i for i, p in enumerate(old_products)}
        for np in new_products:
            pid = np['id']
            if pid in seen_ids:
                # Update existing
                old_products[seen_ids[pid]] = np
            else:
                old_products.append(np)
        generate_landing_links(old_products, write_back=True)
        return jsonify({'ok': True, 'total': len(old_products)})
    return jsonify({'ok': False, 'message': 'No new products'})

@app.route('/api/update_product', methods=['POST'])
def api_update_product():
    if request.headers.get('Authorization') != f"Bearer {ADMIN_TOKEN}":
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    pid = data.get('id')
    if not pid:
        return jsonify({'error': 'Product ID required'}), 400

    products, path = load_catalog()
    for p in products:
        if str(p.get('id')) == str(pid):
            for k in ['title', 'price', 'shipping_price', 'free_shipping']:
                if k in data:
                    p[k] = data[k]
            save_catalog_products(products)
            return jsonify({'ok': True})
    return jsonify({'error': 'Product not found'})

@app.route('/api/archive_orders', methods=['POST'])
def api_archive_orders():
    if request.headers.get('Authorization') != f"Bearer {ADMIN_TOKEN}":
        return jsonify({'error': 'Unauthorized'}), 401

    success, archive_path = archive_and_reset_orders()
    if success:
        # Reset orders.json too? or just Excel sheet
        with open('data/orders.json', 'w', encoding='utf-8') as f:
            json.dump([], f)
        return jsonify({'ok': True, 'archive_path': archive_path})
    return jsonify({'error': 'Archive failed'})

if __name__ == '__main__':
    load_city_area_dict()
    products, _ = load_catalog()
    if not products:
        print("No products found, you may need to update catalog via /api/update_catalog")
    app.run(host='0.0.0.0', port=5000, debug=True)
